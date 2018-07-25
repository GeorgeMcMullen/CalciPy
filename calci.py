#
# CalciPy (calci.py) - A python script for processing ratiometric calcium fluorescence data
#
# By George McMullen
#
# One of the primary challenges in processing ratiometric data is that the two
# signals which make up the ratiometric data are actually captured separately,
# and it is not known which is which. CalciPy processes interlaced calcium
# fluorescence wavelength data, identifies the proper ratio sequence, and
# calculates metrics from the resulting signal. See below for more detail.
# While there is a variety of software for processing calcium fluorescence
# data, each of them have one or more shortcomings, such as:
#
# - Few solutions actually process ratiometric data provided by two
#   wavelengths of light
# - Some require manual effort to determine the proper ratio and
#   subsequent calculations
# - Those that can process ratiometric data automatically are limited
#   (e.g. only able to process a single cell line at a time)
# - Finally, they are not widely available, are closed source, or otherwise
#   proprietary limiting the auditability of the experiments
#
# CalciPy's goal is to enable high throughput, yet customizable, processing
# of ratiometric calcium fluorescence data, in an open and autidible fashion.
# It can also serve as a starting point for processing data that is exported
# from microscopy imaging software, such as Nikon's NIS-Elements.
#


#
#
# Imported libraries (part 1)
#
# Only the basic libraries required by command line arguments are imported first
# in order to speed up execution and error checking when an incorrect argument
# has been used. See "Imported libraries (part 2)" for the rest.
#
#
import argparse
import os
import string

#
#
# Functions
#
#

#
# Argument validation
#
def positiveInteger(string):
    value = int(string)
    if value < 0:
       msg = "%r is not a positive integer" % string
       raise argparse.ArgumentTypeError(msg)
    return value

def positiveFloat(string):
    value = float(string)
    if value < 0:
       msg = "%r is not a positive float" % string
       raise argparse.ArgumentTypeError(msg)
    return value

def floatInRange(string):
    value = float(string)
    if value < 0 or value > 1:
       msg = "%r is not a float between 0.0 and 1.0" % string
       raise argparse.ArgumentTypeError(msg)
    return value

def existingFile(string):
    if os.path.isfile(string) != True:
       msg = "%r file not found" % string
       raise argparse.ArgumentTypeError(msg)
    return string

def existingDir(string):
    if os.path.isdir(string) != True:
       msg = "%r output directory not found" % string
       raise argparse.ArgumentTypeError(msg)
    return string

#
# getMinMax(dataArray)
#
# This is just a wrapper function for peak detection. Other peak detection functions
# can be put in place here without modifying the rest of the code.
#
def getMinMax(dataArray):
    maxima = []
    minima = []

    peaksMax, peaksMin = peakdetect(dataArray, lookahead=peakDetectLookAhead, delta=peakDetectDelta)
    maxima = [p[0] for p in peaksMax]
    minima = [p[0] for p in peaksMin]

    # If the first minima is before the first maxima, then we need to shift it over
    # so we are always start with a max peak, as we will be looking for decay information.
    while len(minima) > 0 and len(maxima) > 0 and minima[0] < maxima[0]:
        minima = minima[1:]

    # We need the maxima and minima to be the same size to we don't get any out of bounds errors
    while (len(minima) > len(maxima)):
        minima = minima[:-1]

    while (len(minima) < len(maxima)):
        maxima = maxima[:-1]

    return [minima, maxima]

#
# getRiseTimes(timeArray, minima, maxima)
#
# How long does it take for the waveform to get from minima (low) to maxima (peak),
# calculating the rise time by explicitly iterating through each minima/maxima.
#
def getRiseTimes(timeArray, minima, maxima):
    riseTimes = []
    for j in range(len(maxima)-1):
        riseTimes = numpy.append(riseTimes, timeArray[maxima[j+1]] - timeArray[minima[j]])

    if len(riseTimes) == 0:
      riseTimes = numpy.append(riseTimes,0)

    return riseTimes

#
# getDecayTimes(timeArray, minima, maxima)
#
# How long does it take for the waveform to get from maxima (peak) to minima (low),
# calculating the decay time by explicitly iterating through each maxima/minima.
#
def getDecayTimes(timeArray, minima, maxima):
    decayTimes = []
    for j in range(len(maxima)):
        decayTimes = numpy.append(decayTimes, timeArray[minima[j]] - timeArray[maxima[j]])

    if len(decayTimes) == 0:
      decayTimes = numpy.append(decayTimes, 0)

    return decayTimes

#
# chooseRatio(dataArray, timeArray)
#
# This is a generic method for choosing the ratio, based on the command line
# argument. It will merely call the other functions.
#
def chooseRatio(dataArray, timeArray):
    if chooseRatioBy == 'time':
       return chooseRatioByTime(dataArray, timeArray)
    else:
       return chooseRatioByAmplitude(dataArray, timeArray)

#
# chooseRatioByTime(dataArray, timeArray)
#
# Since the data contains two different ratios, we need to choose which is the correct one to use.
# The correct ratio signal should have a fast upstroke and slower, exponential decay.
# This function will check to see if the rise is faster, on average, than the decay.
#
# This is not the only method that can do this. Another method would be to check to
# see if there are are more datapoints at the lower end of the amplitude vs. the
# higher end. That is not currently implemented.
#
def chooseRatioByTime(dataArray, timeArray):
    # Create an array which will store the ratios between the two
    # different nm frequencies of light, which are on every other row
    # in each column. It is not known which row is which.
    # The ratio array starts out as being half as tall and twice as wide
    # as the original data array because the data is no longer interlaced.
    testRatioArray1 = dataArray[0::2] / dataArray[1::2]
    testRatioArray2 = dataArray[1::2] / dataArray[0::2]

    # Use the peak detection routing to get maxima and minima, which will then be used to calculate the rise/decay time
    minima, maxima = getMinMax(testRatioArray1)

    currentRangeLimit=rangeLimit
    # Test to make sure that the rangeLimit is within the proper bounds
    if currentRangeLimit < 0 or currentRangeLimit is None or currentRangeLimit > len(minima):
       currentRangeLimit = len(minima)

    riseTimes = getRiseTimes(timeArray, minima, maxima)
    riseTime = riseTimes[:currentRangeLimit].mean()

    decayTimes = getDecayTimes(timeArray, minima, maxima)
    decayTime = decayTimes[:currentRangeLimit].mean()

    if (riseTime < decayTime):
        if invertWaveForm == False:
            ratio = testRatioArray1
        else:
            ratio = testRatioArray2
    else:
        if invertWaveForm == False:
            ratio = testRatioArray2
        else:
            ratio = testRatioArray1

    return ratio

#
# chooseRatioByAmplitude(dataArray, timeArray)
#
# Data that has a sharp excitation phase and an exponential decay phase should
# have more data points residing towards the minimum, and less at the maximum.
# We can use this information to determine which ratio to use.
# 
# TODO: This function could be refinemd so that it looks at individual sets of wavelets
#
def chooseRatioByAmplitude(dataArray, timeArray):
    # Create an array which will store the ratios between the two
    # different nm frequencies of light, which are on every other row
    # in each column. It is not known which row is which.
    # The ratio array starts out as being half as tall and twice as wide
    # as the original data array because the data is no longer interlaced.
    # TODO: We will need to make sure that there is an even number of rows or this step will fail
    testRatioArray1 = dataArray[0::2] / dataArray[1::2]
    testRatioArray2 = dataArray[1::2] / dataArray[0::2]

    ratioMax = max(testRatioArray1)
    ratioMin = min(testRatioArray1)
    
    if sum(1 for i in testRatioArray1 if numpy.abs(i-ratioMin) < numpy.abs(i-ratioMax)) > (len(testRatioArray1) / 2.0):
       if invertWaveForm == False:
          return testRatioArray1
       else:
          return testRatioArray2
    else:
       if invertWaveForm == False:
          return testRatioArray2
       else:
          return testRatioArray1



#
# smoothBetweenPeaks(dataArray, minima, maxima)
#
# This function will take a signal and apply a centered moving average on it, but only between
# the established peaks. This allows the peak values and positions to remain the same, while
# reducing offsets that moving averages can introduce when applied to the entire signal as a whole.
#
def smoothBetweenPeaks(dataArray, minima, maxima):
    smoothedRatio = copy.copy(dataArray)

    for j in range(len(maxima)-1):
      start = minima[j]
      end = maxima[j+1]+1
      if end >= len(dataArray):
        end = maxima[j]
      smoothedRatio[start:end+1] = centeredMovingAverage(dataArray[start:end+1], 5, 2)
      smoothedRatio[start]=dataArray[start]
      smoothedRatio[end]=dataArray[end]

    for j in range(len(minima)):
      start = maxima[j]
      end = minima[j]+1
      if end >= len(dataArray):
        end = minima[j]
      smoothedRatio[start:end+1] = centeredMovingAverage(dataArray[start:end+1], 5, 2)
      smoothedRatio[start]=dataArray[start]
      smoothedRatio[end]=dataArray[end]

    return smoothedRatio

#
# centeredMovingAverage(dataArray, windowSize, numberOfIterations)
#
# As the name implies, this function applies a centered moving average to a signal,
# while at the same time tries to maintain the values found at the start and end of the signal.
#
def centeredMovingAverage(dataArray, windowSize, numberOfIterations):
    smoothedDataArray = copy.copy(dataArray)
    halfWindowSize= int(numpy.floor(windowSize / 2)) # Number of steps on each side to average by
    for q in range(numberOfIterations):
      tempDataArray = copy.copy(smoothedDataArray)
      for k in range(1, len(tempDataArray)-1): # We want to keep the first and last data points the same
        windowStart = k-halfWindowSize
        windowEnd = k+halfWindowSize
        if k < halfWindowSize:
          windowStart=0
        if k > len(dataArray) - halfWindowSize:
          windowEnd=len(dataArray)
        smoothedDataArray[k] = tempDataArray[windowStart:windowEnd].mean()
    return smoothedDataArray

#
# listOfDictionariesMean(listOfDictionaries, key, rangeLimit)
#
# The output of the curve fitting function is a list of dictionaries, with
# the values for Y0, RC, and A contained within. This function will
# calculate the average for a given key within a dictionary.
#
def listOfDictionariesMean(listOfDictionaries, key, rangeLimit):
    if rangeLimit == 0:
       return 0

    currentRangeLimit=rangeLimit
    # Test to make sure that the rangeLimit is within the proper bounds
    if currentRangeLimit < 0 or currentRangeLimit is None or currentRangeLimit > len(listOfDictionaries):
       currentRangeLimit=len(listOfDictionaries)

    totalVal=0
    numberOfValues=0
    for q in range(currentRangeLimit):
      if numpy.isnan(listOfDictionaries[q][key]) == False:
         totalVal += listOfDictionaries[q][key]
         numberOfValues=numberOfValues+1

    if numberOfValues==0:
       return 0

    return totalVal / numberOfValues

#
# getIndexAtAmplitude(dataArray, amplitude, normalize)
#
# This function will return the first index where a signal will reach a given
# amplitude during its decay phase. If normalization is enabled, the signal will be
# normalized by subtracting its lowest value from the entire signal.
# When analyzing calcium fluorescence decay data, amplitude offsets (also referred
# to Calcium Time to Decay or CTD offsets) are used for a variety of reasons,
# including reduction of error with noisy signals.
# Keep in mind that CTD25 means decay to 75% power, and vice versa. CTD75 means decay to 25% power.
# https://www.sciencedirect.com/science/article/pii/S105687191630034X
#
def getIndexAtAmplitude(dataArray, amplitude, normalize):
    if len(dataArray) == 0:
       return 0

    # We make a copy of it because we may normalize the data
    dataArrayCopy = copy.copy(dataArray)

    if (normalize):
       dataArrayCopy = numpy.nan_to_num(dataArrayCopy) - min(numpy.nan_to_num(dataArrayCopy))

    searchValue = max(dataArrayCopy) * amplitude

    closestValueIndex=int((numpy.abs(dataArrayCopy-searchValue)).argmin())

    # Different method which basically just looks for the first point at which the value is less or equal to the amplitude
    #closestValueIndex=0
    #for k in range(len(dataArrayCopy)):
    #  if dataArrayCopy[k] < searchValue:
    #    closestValueIndex=k
    #    break

    return closestValueIndex

#
# getIndicesAtAmplitude(dataArray, minima, maxima, amplitude, normalize)
#
# This function will return the indices where a signal will reach a given
# amplitude during its decay phase.
#
def getIndicesAtAmplitude(dataArray, minima, maxima, amplitude, normalize):
    indices = [];
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        indexAtAmplitude = getIndexAtAmplitude(dataArray[start:end+1], amplitude, normalize)

        indices = numpy.append(indices, int(start+indexAtAmplitude))

    if len(indices) == 0:
      indices = [0]

    return indices

#
# getTimesAtAmplitude(dataArray, timeArray, minima, maxima, amplitude, normalize)
#
# This function will return the time at which a signal reaches a given amplitude
# during its decay phase.
#
def getTimesAtAmplitude(dataArray, timeArray, minima, maxima, amplitude, normalize):
    # Array to store the time when the wave form hits searchValue
    times = []

    # Iterate throughout the entire set of minima/maxima. We'll use the average of all the values later.
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        indexAtAmplitude = getIndexAtAmplitude(dataArray[start:end+1], amplitude, normalize)

        # Using the closest value, we now calculate the time from the maxima to the closest value
        timeAtAmplitude = timeArray[start+indexAtAmplitude]

        times = numpy.append(times, timeAtAmplitude)

    if len(times) == 0:
      times = [0]

    return times

#
# getNormalizedTimesAtAmplitude(dataArray, timeArray, minima, maxima, amplitude, normalize)
#
# This function will return the time at which a signal reaches a given amplitude
# during its decay phase, subtracted by the time that the signal was at it's peak.
#
def getNormalizedTimesAtAmplitude(dataArray, timeArray, minima, maxima, amplitude, normalize):
    # Array to store the time when the wave form hits searchValue
    times = []

    # Iterate throughout the entire set of minima/maxima. We'll use the average of all the values later.
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        indexAtAmplitude = getIndexAtAmplitude(dataArray[start:end+1], amplitude, normalize)

        # Using the closest value, we now calculate the time from the maxima to the closest value
        timeAtAmplitude = timeArray[start+indexAtAmplitude] - timeArray[start]

        times = numpy.append(times, timeAtAmplitude)

    if len(times) == 0:
      times = [0]

    return times

#
# getValuesAtAmplitude(dataArray, minima, maxima, amplitude, normalize)
#
# The above functions take a relative amplitude (0-100%) as their argument.
# This function will return the actual value that is stored in the array
# at the point where the signal reaches a given relative amplitude. While
# one could just calculate the percentage from the max, this will provide
# the actual value, and not just a calculated one.
#
def getValuesAtAmplitude(dataArray, minima, maxima, amplitude, normalize):
    # Array to store the time when the wave form hits searchValue
    valuesArray = []

    # Iterate throughout the entire set of minima/maxima. We'll use the average of all the values later.
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        maxValue=dataArray[start]
        minValue=dataArray[end]

        if (normalize):
           maxValue = numpy.nan_to_num(maxValue) - numpy.nan_to_num(minValue)

        valueAtAmplitude=numpy.nan_to_num(maxValue * amplitude)

        valuesArray = numpy.append(valuesArray, valueAtAmplitude)

    if len(valuesArray) == 0:
      valuesArray = [0]

    return valuesArray

#
# exponentialDecay(x, a, rc, y0)
#
# This function is used to determine the curve of the signal, as well as to plot it.
#
# For more information see: https://en.wikipedia.org/wiki/Exponential_decay#Mean_lifetime
#
# a = The value at the start of the curve
# rc = The decay rate of the curve
# y0 = The baseline for the entire curve. Typically, the last value at the end of the curve. In a normalized curve this would be 0, but for ratiometric calcium fluorescence, the baseline (minimum fluorescence) is typically above 0.
# x = current time/x-axis, typically the X value of the curve
#
def exponentialDecay(x, a, rc, y0):
    return a * numpy.exp(-rc * x) + y0

#
# fitExponentialDecayCurve(dataArray, timeArray, minima, maxima, normalizedDecayTime36)
#
# This function will iterate through an array and use python's curve_fit routine
# determine the exponential decay parameters for each decaying wavelet in the signal.
#
def fitExponentialDecayCurve(dataArray, timeArray, minima, maxima, normalizedDecayTime36):
    exponentialDecayParams = [];
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        # Some error checking because of noisy traces
        if end-start > 6 and end < len(timeArray):
          # Take a slice of the time data and set it to start at 0
          timeSlice = copy.copy(timeArray[start:end+1]) - timeArray[start]

          # Set the best guess parameters based on the existing data and then run the curve fitting
          bestGuess = [dataArray[start] - dataArray[end], 1/normalizedDecayTime36[j], dataArray[end]]

          # Results from curve_fit is often compared to Excel's Solver add-on.
          # TODO: Make bounds parameters as input to these calculations
          lowerBounds = [0.0, 0.0, 0.0]
          upperBounds = [dataArray[start] * 1.4, bestGuess[1]*10, dataArray[end] * 1.4]

          # Use curve_fit to determine the exponential decay parameters given our best guess parameters
          popt=bestGuess
          pcov=[]
          try:
            if useBounds:
               popt, pcov = curve_fit(exponentialDecay, timeSlice, dataArray[start:end+1], bestGuess, maxfev=20000, bounds=[lowerBounds, upperBounds])
            else:
               popt, pcov = curve_fit(exponentialDecay, timeSlice, dataArray[start:end+1], bestGuess, maxfev=20000)
          except RuntimeError:
            print("Error - curve_fit failed")
          except ValueError:
            print("Error - curve_fit failed")

          a, rc, y0 = popt

          if printVerbose == True:
             timeSliceString = ','.join(map(str, timeSlice))
             dataArrayString = ','.join(map(str, dataArray[start:end+1]))
             print "timeSliceString: {0}".format(timeSliceString)
             print "dataArrayString: {0}".format(dataArrayString)
             print "bestGuess: {0}".format(bestGuess)
             print "lowerBounds: {0}".format(lowerBounds)
             print "upperBounds: {0}".format(upperBounds)
             print "a: {0}".format(a)
             print "rc: {0}".format(rc)
             print "y0: {0}".format(y0)
             print "tau: {0}".format(1/rc)
             print
             print


          exponentialDecayParams.append({"A": a, "rc": rc, "y0": y0});
        else:
          exponentialDecayParams.append({"A": numpy.NaN, "rc": numpy.NaN, "y0": numpy.NaN});

    return exponentialDecayParams


#
# getExponentialGoodnessOfFit(dataArray, decayFitData, minima, maxima)
#
# Once the exponential decay curve has been calculated, we can compare it to the
# original data to understand how well they fit as a useful diagnostic measure.
#
def getExponentialGoodnessOfFit(dataArray, decayFitData, minima, maxima):
    goodnessOfFit = []
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        # Some error checking because of noisy traces
        if end-start > 6 and end < len(dataArray):
          # Residual sum of squares
          residualSumOfSquares = numpy.sum((dataArray[start:end+1] - decayFitData[start:end+1]) ** 2)

          # Total sum of squares
          totalSumOfSquares = numpy.sum((dataArray[start:end+1] - numpy.mean(decayFitData[start:end+1])) ** 2)

          # R-Squared
          rSquared = 1 - (residualSumOfSquares / totalSumOfSquares)

          goodnessOfFit = numpy.append(goodnessOfFit, rSquared)
        else:
          goodnessOfFit = numpy.append(goodnessOfFit, 0)

    if len(goodnessOfFit) == 0:
      goodnessOfFit = [0]

    if printVerbose == True:
       goodnessOfFitString = ','.join(map(str, goodnessOfFit))
       print "timeSliceString: {0}".format(goodnessOfFitString)

    return goodnessOfFit


#
# plotExponentialDecayCurve(timeArray, minima, maxima, decayFitData)
#
# Once the parameters for the exponential decay are determined, this function will
# calculate the actual plot points for the curve, using the fit parameters and the
# original exponential decay function.
#
def plotExponentialDecayCurve(timeArray, minima, maxima, decayFitData):
    plottedCurveArray = [numpy.NaN] * len(timeArray)
    for j in range(len(minima)):
        start = int(maxima[j])
        end = int(minima[j])

        # Some error checking because of noisy traces
        if end-start > 6 and end < len(timeArray):
          # Take a slice of the time data and set it to start at 0
          timeSlice = copy.copy(timeArray[start:end+1]) - timeArray[start]

          # Create a new curve based on the same exponential decay function, using the found parameters from curve_fit
          a = decayFitData[j]["A"]
          rc = decayFitData[j]["rc"]
          y0 = decayFitData[j]["y0"]
          plottedCurveArray[start:end+1] = exponentialDecay(timeSlice, a, rc, y0)

    return plottedCurveArray



#
#
# Argument parsing
#
#

argumentParser = argparse.ArgumentParser(prog='calci.py', description='Process ratiometric calcium fluorescence decay data.')

argumentParser.add_argument('inputfile', metavar='filename', type=existingFile, help='input filename to process')

argumentParser.add_argument('-O', '--outputdir', type=existingDir, default='.', help='output files to a directory')

argumentParser.add_argument('--sheetname', help='process only a specific worksheet by name (use quotes)')
argumentParser.add_argument('--sheetnum', type=positiveInteger, help='process only a specific worksheet by number')
argumentParser.add_argument('--column', type=positiveInteger, help='process only a specific cell line column of data (typically 1-10)')

argumentParser.add_argument('--bgreduce', choices=['average', 'point', 'none'], default='average',  help='method for reducing background data')

argumentParser.add_argument('--lookahead', type=positiveInteger, default=30, help='peak detection look-ahead parameter (default=30)')
argumentParser.add_argument('--delta',     type=positiveFloat, default=0.0,  help='peak detection delta parameter (default=0)')

argumentParser.add_argument('--invert', action='store_true', help='invert the waveform (special cases only)')
argumentParser.add_argument('--ratioby', choices=['time', 'amplitude'], default='time',  help='method for choosing which ratio to use')

argumentParser.add_argument('--decaystart', type=floatInRange, default=1.0, help='the relative amplitude from the peak where the analysis will start')
argumentParser.add_argument('--decayend',   type=floatInRange, default=0.0, help='the relative amplitude from the peak where the analysis will end')

argumentParser.add_argument('--bounds', action='store_true', help='bound the solution to non-negative numbers, around the start/end values of the waveform')

argumentParser.add_argument('--limit', type=positiveInteger, help='limit processing to the first X wavelets (default=0 as disabled)')

argumentParser.add_argument('--ymax', type=float, help='Y-axis maximum')
argumentParser.add_argument('--ymin', type=float, help='Y-axis minimum')

argumentParser.add_argument('--xmax', type=float, help='X-axis maximum')
argumentParser.add_argument('--xmin', type=float, help='X-axis minimum')

argumentParser.add_argument('--show', action='store_true', help='show the graph (with matplotlib)')

argumentParser.add_argument('--verbose', action='store_true', help='print values as they are calculated')

#
#
# Variable initialization
#
#

commandArguments = argumentParser.parse_args()

# inputFileName: Name of the file to process
inputFileName = commandArguments.inputfile

# outputDirectory: The location where output files will be stored
outputDirectory = commandArguments.outputdir

# inputFileNameNoExt: The file's name without any extension
inputFileNameNoExt = os.path.basename(inputFileName).replace(".xlsx","")
inputFileNameNoExt = inputFileNameNoExt.replace(".xls","")
inputFileNameNoExt = inputFileNameNoExt.replace(".csv","")

# outputCSVFileName: The name of the output CSV file
outputCSVFileName = os.path.join(outputDirectory,inputFileNameNoExt+'_Processed.csv')

worksheetToProcess = commandArguments.sheetname
worksheetNumToProcess = commandArguments.sheetnum
if worksheetToProcess is not None and worksheetNumToProcess is not None:
   print "error: argument --sheetname cannot be used at the same time as --sheetnum"
   quit(-1)


# columnToProcess: A specific column to process (depends on worksheetToProcess / worksheetNumToProcess)
columnToProcess = None
if commandArguments.column is not None:
   if worksheetToProcess is not None or worksheetNumToProcess is not None:
      columnToProcess = commandArguments.column
   else:
      print "error: argument --column requires either argument --sheetname or --sheetnum"
      quit(-1)

backgroundReduceMode = commandArguments.bgreduce

peakDetectLookAhead = commandArguments.lookahead
peakDetectDelta = commandArguments.delta

invertWaveForm = commandArguments.invert
chooseRatioBy = commandArguments.ratioby

rangeLimit = commandArguments.limit

useBounds = commandArguments.bounds

# The relative amplitude from the peak where the analysis will start and end.
# This is the inverse of the CTD, which indicates how much the signal has decayed by.
decayStart = commandArguments.decaystart
decayEnd = commandArguments.decayend

if decayStart <= decayEnd:
   print "error: argument --decaystart should be greater than argument --decayend"
   quit(-1)

graphYMax = commandArguments.ymax
graphYMin = commandArguments.ymin

if graphYMax is not None and graphYMin is not None and graphYMax <= graphYMin:
   print "error: argument --ymax should be greater than argument --ymin"
   quit(-1)

graphXMax = commandArguments.xmax
graphXMin = commandArguments.xmin

if graphXMax is not None and graphXMin is not None and graphXMax <= graphXMin:
   print "error: argument --xmax should be greater than argument --graphXMin"
   quit(-1)

# showGraph: Use matplotlib to display the graphs
showGraph = commandArguments.show

# printVerbose: Print out values as they are calculated
printVerbose = commandArguments.verbose

# capitalLetters: All capital letters, which will be used to get the cells from an Excel File
capitalLetters = string.ascii_uppercase

#
#
# Imported libraries part 2
#
# Some libraries, like matplotlib, seem to execute code when getting imported.
# So saving this step for at least after the command line arguments are parsed.
#
import math
import copy
from openpyxl import load_workbook         # pip install openpyxl
import numpy as numpy
from scipy.optimize import curve_fit       # pip install scipy
#from peakdetect import peakdetect          # https://gist.github.com/sixtenbe/1178136
from analytic_wfm import peakdetect        # pip install analytic-wfm
import matplotlib.pyplot as pyplot         # pip install matplotlib
import matplotlib.gridspec as gridspec



#
#
# Main Program
#
#

#
# Load the inputFileName Excel file
#
excelFile = load_workbook(filename = inputFileName)
allWorkSheetNames = excelFile.sheetnames

# If we have specified a worksheet, make sure it can be found
if worksheetToProcess:
   if allWorkSheetNames.index(worksheetToProcess):
      allWorkSheetNames=[worksheetToProcess]
   else:
      print "error: argument --sheetname worksheet "+worksheetToProcess+" does not exist"
      quit(-1)
elif worksheetNumToProcess is not None:
   if worksheetNumToProcess < len(allWorkSheetNames):
      allWorkSheetNames=[allWorkSheetNames[worksheetNumToProcess]]
   else:
      print "error: argument --sheetnum worksheet "+worksheetNumToProcess+" does not exist"
      quit(-1)

# Open the CSV for writing and write the header
outputCSVFile = open(outputCSVFileName, 'w')
outputCSVFile.write('Sheet,')
outputCSVFile.write('Cell Line Name,')
outputCSVFile.write('Cell Line #,')
outputCSVFile.write('Min,')
outputCSVFile.write('Max,')
outputCSVFile.write('Amplitude,')
outputCSVFile.write('Beat Rate,')
outputCSVFile.write('Beat Variation,')
outputCSVFile.write('Rise Time,')
outputCSVFile.write('Rise Velocity,')
outputCSVFile.write('Y0,')
outputCSVFile.write('A,')
outputCSVFile.write('RC,')
outputCSVFile.write('Tau,')
outputCSVFile.write('RC1d,')
outputCSVFile.write('RC2d,')
outputCSVFile.write('Goodness of Fit,')
outputCSVFile.write('Decay Time,')
outputCSVFile.write('Curve Fit Decay Time 36%,')
outputCSVFile.write('Curve Fit Decay Value 36%,')
outputCSVFile.write('\n')

# Loop through the worksheets in the given Excel file and process them
for workSheetName in allWorkSheetNames:
    # Load the Excel work sheet
    # Deprecated function get_sheet_by_name. Use wb[sheetname]
    #workSheet = excelFile.get_sheet_by_name(workSheetName)
    workSheet = excelFile[workSheetName]

    print 'Processing: '+inputFileName+' - '+workSheetName

    # Sometimes we get additional columns with notes in them. That will crash the application.
    # So we manually limit to 16 columns (value of 15 with 0 indexing)
    workSheetMaxColumn = workSheet.max_column

    # If we don't have enough columns in the sheet, then skip it
    if (workSheetMaxColumn < 7):
        continue

    # Just limit the number of columns to the P column
    if (workSheetMaxColumn > 16):
        workSheetMaxColumn=16

    # Create array for the data. The typical sheet in the Excel file has
    # one header row and five columns of additional informaion like time.
    # It may also have up to ten individual columns of data.
    dataArray = numpy.ndarray(shape=(workSheet.max_row-1, workSheetMaxColumn-5))

    # Create an array that will contain the column headers
    columnHeaders = []

    # Create an array for the time signatures of each row in the work sheet.
    timeArray = numpy.ndarray(shape=(workSheet.max_row-1, 1))

    # Fill the data arrays with data from the current work sheet
    for i in range(5,workSheetMaxColumn):
        for j in range(2,workSheet.max_row+1):
            dataArray[j-2, i-5] = workSheet[capitalLetters[i] + str(j)].value

    # Fill the data arrays with data from the current work sheet
    for i in range(5,workSheetMaxColumn):
        columnHeaders.append(workSheet[capitalLetters[i] + str(1)].value)

    # Fill the time array with data from the current work sheet
    for j in range(2,workSheet.max_row+1):
        timeArray[j-2] = workSheet[capitalLetters[1] + str(j)].value

    # Background data is typically the last column in each work sheet
    # We need to separate it out into its own array and remove the background
    # noise from the actual data.
    backGroundDataArray = dataArray[:,-1]
    dataArray = dataArray[:,:-1]

    for q in range(workSheetMaxColumn-6):
      if backgroundReduceMode == 'average':
         dataArray[::2,q] = dataArray[::2,q] - backGroundDataArray[::2].mean()
         dataArray[1::2,q] = dataArray[1::2,q] - backGroundDataArray[1::2].mean()
      elif backgroundReduceMode == 'point':
         dataArray[:,q] = dataArray[:,q] - backGroundDataArray

    # We need to make sure that there is an even number of rows or the script will fail when choosing a ratio with two differently sized arrays.
    if len(timeArray) % 2 != 0:
      dataArray=dataArray[:-1,:]
      timeArray=timeArray[:-1]

    # Once we choose the proper ratios, the size of the working data will be half as tall,
    # we also need to do the same with the time array as they should be the same size.
    # Otherwise calculations for BPM, tau, rise time, etc. will be off.
    timeArray = timeArray[::2]
    timeArray = timeArray.flatten()

    if columnToProcess is None:
       columnsToProcess = range(dataArray.shape[1])
    else:
       if columnToProcess < workSheetMaxColumn-5:
          columnsToProcess = [columnToProcess]
       else:
          columnsToProcess = range(dataArray.shape[1])

    #
    # Configure the graph
    #

    # Close all figures as a precaution
    pyplot.close('all')

    # Starts a new figure
    if showGraph == True:
       figure = pyplot.figure()
    else:
       figure = pyplot.figure(figsize=(19.20,10.80))

    pyplot.suptitle(inputFileName+' - '+workSheetName)

    if len(columnsToProcess) == 1:
       # Create a "grid" of just one row and column
       currentGridSpec = gridspec.GridSpec(1,1)
    else:
       # Create a new grid for the plots for two columns (and half the number of rows)
       currentGridSpec = gridspec.GridSpec(int(math.ceil(len(columnsToProcess)/2.0)), 2)

    currentGridPosition=0

    # Loop through each of the newly calculated ratio columns
    for i in columnsToProcess:
        # We take the data for the current column, calculate ratios and choose the proper ratio
        ratioArray=chooseRatio(dataArray[:,i], timeArray)

        # Get the peaks of the waveform as minima and maxima positions
        minima, maxima = getMinMax(ratioArray)

        # Test to make sure that the rangeLimit is within the proper bounds
        currentRangeLimit = rangeLimit
        if currentRangeLimit < 0 or currentRangeLimit is None or currentRangeLimit > len(minima):
          currentRangeLimit = len(minima)

        # In order to get a decent first estimate for curve fitting, we are going to smooth
        # the wave form, but only between the peaks that we've found. That way it will not get
        # offset by the rise signal or areas where the signal crests around the peaks.
        smoothedRatioArray = smoothBetweenPeaks(ratioArray, minima, maxima)

        offsetMaxima = getIndicesAtAmplitude(smoothedRatioArray, minima, maxima, decayStart, True)
        offsetMinima = getIndicesAtAmplitude(smoothedRatioArray, minima, maxima, decayEnd, True)

        minimaMean = ratioArray[minima[:currentRangeLimit]].mean()
        maximaMean = ratioArray[maxima[:currentRangeLimit]].mean()

        amplitude = maximaMean - minimaMean

        beatRate = 0
        beatVariation = 0
        if len(maxima) > 1:
          beatInterval = numpy.diff(timeArray[maxima]).mean()
          if beatInterval != 0:
            beatRate = (1. / beatInterval * 60) # in beats/minute
            beatVariation = numpy.std((1. / numpy.diff(timeArray[maxima])))

        riseTimesArray = getRiseTimes(timeArray, minima, maxima)
        riseTime = riseTimesArray[:currentRangeLimit].mean()
        #riseTime = riseTimesArray.mean()
        riseVelocity = 0
        if riseTime !=0:
          riseVelocity = amplitude / riseTime

        decayTime36 = getNormalizedTimesAtAmplitude(smoothedRatioArray, timeArray, offsetMinima, offsetMaxima, 0.367879441171442, True)

        decayTimesArray = getDecayTimes(timeArray, minima, maxima)
        decayTime = decayTimesArray[:currentRangeLimit].mean()
        #decayTime = decayTimesArray.mean()

        decayFitData = fitExponentialDecayCurve(ratioArray, timeArray, offsetMinima, offsetMaxima, decayTime36)
        decayCurveData = plotExponentialDecayCurve(timeArray, offsetMinima, offsetMaxima, decayFitData)
        goodnessOfFit = getExponentialGoodnessOfFit(ratioArray, decayCurveData, offsetMinima, offsetMaxima)

        tauMarkerArray=getIndicesAtAmplitude(decayCurveData, offsetMinima, offsetMaxima, 0.367879441171442, True)
        tauValueArray=getValuesAtAmplitude(decayCurveData, offsetMinima, offsetMaxima, 0.367879441171442, True)

        # Calculate the correct tau (should also be able to get this data from the popt)
        curveFitDecayTime36 = getNormalizedTimesAtAmplitude(decayCurveData, timeArray, offsetMinima, offsetMaxima, 0.367879441171442, True)

        tau=0
        rc=listOfDictionariesMean(decayFitData, "rc", currentRangeLimit)
        if rc != 0:
          tau = 1 / rc

        rc1d = rc * 3
        rc2d = rc / 2

        #
        # Plotting of Graph
        #

        # Add a new chart to the grid
        chartSubPlot = figure.add_subplot(currentGridSpec[currentGridPosition])
        currentGridPosition = currentGridPosition + 1

        pyplot.title(columnHeaders[i])

        # Set the min and max for the Y axis
        if graphYMax is not None:
           pyplot.ylim(top=graphYMax)
        if graphYMin is not None:
           pyplot.ylim(bottom=graphYMin)

        # Set the min and max for the X axis
        if graphXMax is not None:
           pyplot.xlim(right=graphXMax)
        else:
           pyplot.xlim(right=timeArray[-1])

        if graphXMin is not None:
           pyplot.xlim(left=graphXMin)
        else:
           pyplot.xlim(left=timeArray[0])


        # Chosen 340/380nm ratio waveform
        chartSubPlot.plot(timeArray, ratioArray, 'm')

        # Fitted decay curves
        chartSubPlot.plot(timeArray, decayCurveData, 'b')

        # Marker for where the fit curve hits 1/e
        for tauMarker in tauMarkerArray:
            chartSubPlot.plot([timeArray[int(tauMarker)],timeArray[int(tauMarker)]], [decayCurveData[int(tauMarker)]-(amplitude*.2),decayCurveData[int(tauMarker)]+(amplitude*.2)], 'r')

        # Peak maximum data points as yellow dots
        for maximum in maxima:
            chartSubPlot.plot(timeArray[maximum], ratioArray[maximum], 'yo')

        # Peak minimum data points as green dots
        for minimum in minima:
            chartSubPlot.plot(timeArray[minimum], ratioArray[minimum], 'go')

        #
        # Output of data to CSV file
        #
        outputCSVFile.write(workSheetName + ', ')
        outputCSVFile.write(columnHeaders[i] + ', ')
        outputCSVFile.write(str(i+1) + ', ')
        outputCSVFile.write(str(minimaMean) + ', ')
        outputCSVFile.write(str(maximaMean) + ', ')
        outputCSVFile.write(str(amplitude) + ', ')
        outputCSVFile.write(str(beatRate) + ', ')
        outputCSVFile.write(str(beatVariation) + ', ')
        outputCSVFile.write(str(riseTime) + ', ')
        outputCSVFile.write(str(riseVelocity) + ', ')
        outputCSVFile.write(str(listOfDictionariesMean(decayFitData, "y0", currentRangeLimit)) + ', ')
        outputCSVFile.write(str(listOfDictionariesMean(decayFitData, "A", currentRangeLimit)) + ', ')
        outputCSVFile.write(str(listOfDictionariesMean(decayFitData, "rc", currentRangeLimit)) + ', ')
        outputCSVFile.write(str(tau) + ', ')
        outputCSVFile.write(str(rc1d) + ', ')
        outputCSVFile.write(str(rc2d) + ', ')
        outputCSVFile.write(str(goodnessOfFit[:currentRangeLimit].mean()) + ', ')
        outputCSVFile.write(str(decayTime) + ', ')
        outputCSVFile.write(str(curveFitDecayTime36[:currentRangeLimit].mean()) + ', ')
        outputCSVFile.write(str(tauValueArray[:currentRangeLimit].mean()) + ', ')
        outputCSVFile.write('\n')

    #
    # Finished processing worksheet
    #

    # Create a new PDF file for each work sheet, which will show the plotted graphs
    outputPDFFileName = os.path.join(outputDirectory,inputFileNameNoExt + '_' + workSheetName + '.pdf')

    # Improve the layout of the grid
    currentGridSpec.tight_layout(figure, rect=[0, 0.03, 1, 0.95], h_pad=0, w_pad=0)

    # Output the plot as a PDF file
    pyplot.savefig(outputPDFFileName, dpi=1200)

    # If we want to view the image generated, uncomment the below file
    if showGraph:
       pyplot.show()

    # Make sure all figures are closed
    pyplot.close('all')

outputCSVFile.close()
